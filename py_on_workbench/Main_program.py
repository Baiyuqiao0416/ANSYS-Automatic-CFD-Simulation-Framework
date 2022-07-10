# -*- coding: utf-8 -*-
# Based on ANSYS 2019R2
# Python 3.7 x64

#  ----------------------------------------------------------------------------------
# |                                                                                  |
# |  Copyright © 2021 by Yuqiao Bai & Xiangyu Zhao of AOT Lab in Soochow University  |
# |                                                                                  |
#  ----------------------------------------------------------------------------------

import os 
import sys
import time
import shutil
import psutil 
import decimal
import datetime
import platform
import linecache
import fileinput
import xml.dom.minidom
import xml.etree.cElementTree
import openpyxl #third package
from  distutils import util
from PyWbUnit import CoWbUnitProcess
from multiprocessing import cpu_count

# Fluent setting
NumberOfProcessors=int(cpu_count())-2
NumberOfGPGPUs='1'

# Work setting
# Use relative paths instead
SCDMscriptname='SCDM_Script.py' 
Meshscriptname='MESH_Script.py'
Fluentscriptname='FLUENT_Script.scm'  
Additerationscriptname='Add_Iteration_Script.scm'

#Log
class Logger(object):

    def __init__(self, stream=sys.stdout):
        output_dir = "log"
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        log_name = '{}.log'.format(time.strftime('%Y-%m-%d-%H-%M'))
        filename = os.path.join(output_dir, log_name)

        self.terminal = stream
        self.log = open(filename, 'a+')

    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)

    def flush(self):
        pass

sys.stdout = Logger(sys.stdout)  # write console to log
sys.stderr = Logger(sys.stderr)  # write err to log 

#cpu usage
def cpu_usage():
    while True:                   
        cpu = psutil.cpu_percent(interval=0.1)   
        return float(cpu)

#xml
def xml_getdata(dir,tagname):
    file = xml.dom.minidom.parse(dir)
    tagtext=file.getElementsByTagName(tagname)
    tagtext1=tagtext[0]
    return tagtext1.firstChild.data.strip('\n')
def xml_changedata(dir,tagname,tagtext):
    updateTree = xml.etree.cElementTree.parse(dir) 
    root = updateTree.getroot()
    tag=root.find(tagname)
    tag.text = tagtext
    updateTree.write(dir)

#last line
def get_last_line(filedir):
    file=open(filedir, 'rb')
    off = -50                     #offset
    while True:
        file.seek(off, 2)            #seek(off, 2)represents the file pointer: 50 characters (-50) forward from the end of the file (2)
        lines = file.readlines()     #Read all lines in the file pointer range
        if len(lines)>=2:         #Determine whether there are at least two lines at the end, so as to ensure that the last line is complete
            last_line = lines[-1] #Take the last line
            break
        #If the readlines obtained when off is 50, there is only one line of content, then there is no guarantee that the last line is complete
        #So the off doubles and reruns until the readlines is more than one line
        off *= 2
    file.close()
    return last_line.split()

#replace specific line
def replacement(file, previousw, nextw):
   for line in fileinput.input(file, inplace=1):
       line = line.replace(previousw, nextw)
       sys.stdout.write(line)

#count the total lines
def count_lines(file):
    count = -1
    for count, line in enumerate(open(file, 'rb')):
        pass
        count += 1
    return count

#if last 1-10 phead average <±5 than the previous 2-11. Compare 10 dp.
def compare_phead(file):
    count = count_lines(file)
    i=0
    p=[0 for x in range(0, 20)]
    while(i<20):
        text1=linecache.getline(file,count-i).split()
        p[i]=float(text1[1])
        i+=1
    m=0
    j=0
    while(j<10):
        k=0
        sum1=sum2=0
        while(j+k<=j+9):
            sum1+=p[k+j]
            sum2+=p[k+j+1]
            k+=1
        avg_dp=(sum1-sum2)/10
        j+=1
        if(avg_dp >= -5 and avg_dp <= 5):
            m+=1
    if(m==10):
        result=(sum1+sum2)/20
        return result  #Pa
    else:
        return '0'

#Determine whether the calculation is complete
def wait_caculation():
    while(True):
        cpu_percent=cpu_usage()
        if(cpu_percent < 90.0 and cpu_percent != 0.0):
            break
        else:
            time.sleep(120)

print()
print('====================================================================================')
print()
print('                             ,@@@@\.       ,/@@@@@@@@`. @@@@@@@@@@@@@@@@@@@@@.      ')
print('                            =@@@@@@@.    ,@@@`.   .\@@@^,@@@@@@@@@@@@@@@@@@@`       ')
print('                           ,@@@@@@@@\   =@@`  .` @@`,@@@^      =@@@@@@.             ')
print('                          .@@@@@@@=@@^  @@^  /@@. ,@^=@@@.     =@@@@@@.             ')
print('                          /@@@@@@. \@@` @@^  @@`  ,@^=@@@.     =@@@@@@.             ')
print('                         =@@@@@@`  .@@@.=@@`   ,\/@/,@@@^      =@@@@@@.             ')
print('    .*******]@@@@\`*.   ,@@@@@@^    ,@@\.,@@\`. ..,@@@@^       =@@@@@@.             ')
print('   [[[[[[[[`.  .*[[\@..@@@@@@^      =@@^  ,\@@@@@@@@^]]*.     =@@@@@@.              ')
print('                    .@\/@@@@@/        \@@` .]]]/@@@@[[**[@@]****\@@@/*********..    ')
print('                      \@@@@@`         .@@@/@`**.          .,[[[[[[`[[[[[[[[[[[[.    ')
print('                                       ,@@/                                         ')
print('                                        ,[                                          ')
print()
print('====================================================================================') 
print()
print('                 SOOCHOW UNIVERSITY ARTIFICIAL ORGAN TECHNOLOGY LAB                 ')                                                                                
print()
print('                         Automatic CFD Simulation Framework                         ')
print()
print('                               Based On ANSYS 2019 R2                               ')
print()
print('====================================================================================')
print()
print('System Information')
print()
print('Platform:'+platform.platform())
print('System Version:'+platform.version())
print('System Architecture:'+str(platform.architecture()))
print('Computer Name:'+platform.node())
print('Processor:'+platform.processor())
print('Number of logical processors:'+str(cpu_count()))
print()
print('====================================================================================')

#xml settings
setfile=sys.path[0]+'//settings.xml'
#get flag
flag=bool(util.strtobool(xml_getdata(setfile,'is_seted')))
if(flag):
    print()
    print('                             Initial setup is completed.                            ')
    print()
    print('====================================================================================')
while(not flag):
    print()
    print('Please define the workbench path:')
    workbench_dir=input()
    #print('Please define the material database path:')
    material_database_dir0=sys.path[0]
    print()
    print('Initial setup is completed.')
    print()
    print('====================================================================================')
    material_database_command='(cx-gui-do cx-set-text-entry "Open Database*TextEntry1(Database Name)" "'+ material_database_dir0.replace('\\','/') +'/blood.scm")' 
    xml_changedata(setfile,'workbench_dir',workbench_dir)
    xml_changedata(setfile,'material_database_command',material_database_command)
    xml_changedata(setfile,'is_seted','True')
    #get flag
    flag=bool(util.strtobool(xml_getdata(setfile,'is_seted')))

# Input parameters
in_content="N"
while(in_content!="Y"):
    print("Please define the work path:")
    workpath = input()
    print('Please set the volume flow (LPM):')
    volume_flow=input()   
    print("Please define the initial d 'd0' (mm):")
    d0 = input()
    print("Please define the delta d 'step' (mm):")
    step = input()
    print("Please define the number of caculation 'n':")
    n = input()
    print()
    print("Please make sure the following content is correct, press 'Y' if it is correct.")
    print()
    print('workpath='+workpath)
    print('volume_flow='+volume_flow+' LPM')
    print('d0='+d0+' mm')
    print('step='+step+' mm')
    print('n='+n)
    print()
    in_content=input()
    
print('====================================================================================')
print()
print('                                  Initializing....                                  ')
print()
print('====================================================================================')


mass_flow_command="(cx-gui-do cx-set-expression-entry \"Mass-Flow Inlet*Frame3*Frame1(Momentum)*Table1*Table8*ExpressionEntry1(Mass Flow Rate)\" \'(\""+str(round(float(float(volume_flow)/60*1.055),5))+"\" . 0))"
xml_changedata(setfile,'mass_flow_command',mass_flow_command)
#get command 
WorkbenchDir = str(xml_getdata(setfile,'workbench_dir'))
MaterialCommand = str(xml_getdata(setfile,'material_database_command'))
MassFlowCommand = str(xml_getdata(setfile,'mass_flow_command'))

#change fluentscript
var1 = "material_database_command"
replacement(Fluentscriptname, var1, MaterialCommand)
var2 = "mass_flow_command"
replacement(Fluentscriptname, var2, MassFlowCommand)

wb=openpyxl.Workbook() #create Workbook() object
ws=wb.active #get dafault sheet
ws.append(["Name", "p_head(Pa)","p_head(mmHg)"])#write data to file
wb.save(workpath+"\\caculate_dp.xlsx")

# loop
i=0
while(i<int(n)):
    try:
        d=decimal.Decimal(d0)+decimal.Decimal(step)*decimal.Decimal(i)
        name=str(i+1)+"_"+str(d)+"mm"
        wbpjname=name+".wbpj "
        if not os.path.exists(workpath+"\\"+name):
            os.mkdir(workpath+"\\"+name)
        
        # Create an instance of the wb unit and specify the ansys wb version
        coWbUnit = CoWbUnitProcess(WorkbenchDir, version=194, interactive=True) 
        coWbUnit.initialize()
        time.sleep(60)
        coWbUnit.execWbCommand('SetScriptVersion(Version="19.4.159")')
        coWbUnit.execWbCommand('system1 = GetTemplate(TemplateName="Fluid Flow").CreateSystem()')

        print()
        print('====================================================================================')
        print()
        print('n='+str(i+1))
        print('['+datetime.datetime.now().strftime('%F %T')+']\0'+name+'\0Workbench Initialized.')
        t0=time.time()

        # SCDM
        coWbUnit.execWbCommand('Geometry1 = system1.GetContainer(ComponentName="Geometry")')
        coWbUnit.execWbCommand('Geometry1.Edit(IsSpaceClaimGeometry=True)')
        SCDMscript=open(SCDMscriptname,"r",encoding='utf-8')
        SCDMcmd='d='+str(d)+'\n'+'NumberOfProcessors='+str(NumberOfProcessors)+'\n'+str(SCDMscript.read())
        coWbUnit.execWbCommand('SCDMcmd="""'+SCDMcmd+'"""')
        coWbUnit.execWbCommand('Geometry1.SendCommand(Language="Python",Command=SCDMcmd)')
        coWbUnit.execWbCommand('Geometry1.Exit()')
        SCDMscript.close()
        print('['+datetime.datetime.now().strftime('%F %T')+']\0'+name+'\0SCDM Completed.')

        # Meshing
        coWbUnit.execWbCommand('MeshComponent1 = system1.GetComponent(Name="Mesh")')
        coWbUnit.execWbCommand('MeshComponent1.Refresh()')
        coWbUnit.execWbCommand('Mesh1 = system1.GetContainer(ComponentName="Mesh")')
        coWbUnit.execWbCommand('Mesh1.Edit()')
        Meshscript=open(Meshscriptname,"r",encoding='utf-8')
        #Sizing of throat <=0.13125mm (0.075d)
        if(d<=1.75 and d>0.75):
            Meshcmd='mesh1 = Model.Mesh'+'\n'+'mesh1.NumberOfCPUsForParallelPartMeshing='+str(NumberOfProcessors)+'\n'+'d='+str(d)+'\n'+str(Meshscript.read())
        elif(d<=0.75):
            Meshcmd='mesh1 = Model.Mesh'+'\n'+'mesh1.NumberOfCPUsForParallelPartMeshing='+str(NumberOfProcessors)+'\n'+'d=0.75'+'\n'+str(Meshscript.read())    
        else:
            Meshcmd='mesh1 = Model.Mesh'+'\n'+'mesh1.NumberOfCPUsForParallelPartMeshing='+str(NumberOfProcessors)+'\n'+'d=1.75'+'\n'+str(Meshscript.read())
        coWbUnit.execWbCommand('Meshcmd="""'+Meshcmd+'"""')
        coWbUnit.execWbCommand('Mesh1.SendCommand(Language="Python",Command=Meshcmd)')
        coWbUnit.execWbCommand('Mesh1.Exit()')
        coWbUnit.execWbCommand('MeshComponent1.Update(AllDependencies=True)')
        Meshscript.close()
        print('['+datetime.datetime.now().strftime('%F %T')+']\0'+name+'\0Meshing Completed.')
        
        try:
            coWbUnit.saveProject(workpath+"\\"+name+"\\"+wbpjname)
            os.mkdir(workpath+"\\"+name+"\\"+name+"_files"+"\\dp0\\FFF\\Fluent")
            # Fluent
            source = "output_residual.jou"
            target = workpath+"\\"+name+"\\"+name+"_files"+"\\dp0\\FFF\\Fluent"
            shutil.copy(source, target)
            source2 = "residuals.dat"
            shutil.copy(source2, target)

            coWbUnit.execWbCommand('setupComponent1 = system1.GetComponent(Name="Setup")')
            coWbUnit.execWbCommand('setupComponent1.Refresh()')
            coWbUnit.execWbCommand('setup1 = system1.GetContainer(ComponentName="Setup")')
            coWbUnit.execWbCommand('fluentLauncherSettings1 = setup1.GetFluentLauncherSettings()')
            coWbUnit.execWbCommand('fluentLauncherSettings1.SetEntityProperties(Properties=Set(Precision="Double", EnvPath={}, RunParallel=True, NumberOfProcessors='+str(NumberOfProcessors)+', NumberOfGPGPUs=1))')
            coWbUnit.execWbCommand('setup1.Edit()')
            Fluentscript=open(Fluentscriptname,"r",encoding='utf-8')
            Fluentline=Fluentscript.readline()
            while Fluentline:
                Fluentline=Fluentline.strip('\n')
                coWbUnit.execWbCommand('setup1.SendCommand("""'+Fluentline+'""")')
                Fluentline=Fluentscript.readline()
            Fluentscript.close()
            #time.sleep(120)
            #if residual > 1e-05 , excute another journal to increase the iteration (up to 10000 iterations)
            
            fname_residuals = workpath+"\\"+name+"\\"+name+"_files"+"\\dp0\\FFF\\Fluent\\residuals.dat"
            fname_p_head = workpath+"\\"+name+"\\"+name+"_files"+"\\dp0\\FFF\\Fluent\\p-head-rfile.out"
            fname_d_mfr = workpath+"\\"+name+"\\"+name+"_files"+"\\dp0\\FFF\\Fluent\\diff-mfr-rfile.out"

            #Determine whether the calculation is complete
            wait_caculation()
            wb=openpyxl.load_workbook(workpath+"\\caculate_dp.xlsx")
            sheet=wb['Sheet'] #Get sheet by name           
            #lines of residuals.dat
            count = -1
            countnext=0
            while(True):
                count = count_lines(fname_residuals)      
                if(countnext==count+1):
                    print('['+datetime.datetime.now().strftime('%F %T')+']\0'+str(count-1)+' iterations completed.')
                    break
                countnext=count+1
            iteration=count-1
            addcount=0
            while(True):
                a=get_last_line(fname_residuals)
                b=get_last_line(fname_d_mfr)
                c=compare_phead(fname_p_head)
                print()
                print('iterations= '+str(iteration))
                print('continuity= {:.4e}'.format(float(a[0])))
                print('x-velocity= {:.4e}'.format(float(a[1])))
                print('y-velocity= {:.4e}'.format(float(a[2])))
                print('z-velocity= {:.4e}'.format(float(a[3])))
                print('         k= {:.4e}'.format(float(a[4])))	
                print('     omega= {:.4e}'.format(float(a[5])))
                print(' delta mfr= {:.4e}'.format(float(b[1])))  #delta mass flow rate
                print('    p-head= {:.4e} Pa'.format(float(c)))
                print('          = {:.4e} mmHg'.format(float(c)*0.0075))
                print()
                if(addcount==4):
                    break
                if(compare_phead(fname_p_head) == '0'):
                    print('Not converged, add 2000 iterations.')
                    Additerationscript=open(Additerationscriptname,"r",encoding='utf-8')
                    Additerationcmd=Additerationscript.readline()
                    while Additerationcmd:
                        Additerationcmd=Additerationcmd.strip('\n')
                        coWbUnit.execWbCommand('setup1.SendCommand(Command="""'+Additerationcmd+'""")')
                        Additerationcmd=Additerationscript.readline()
                    Additerationscript.close()
                    addcount+=2
                    count = -1
                    wait_caculation()
                    while(True):
                        count = count_lines(fname_residuals)   
                        if(countnext==count+1):
                            print('['+datetime.datetime.now().strftime('%F %T')+']\0'+str(count-addcount)+' iterations completed.')
                            break
                        countnext=count+1
                    iteration+=(count-addcount)
                else:
                    print('Converged.')
                    break
                sheet.append([name, round(float(c),4),round(float(c)*0.0075,4)]) 
            #after add iterations still not converged
            if(compare_phead(fname_p_head) == '0'):
                print('['+datetime.datetime.now().strftime('%F %T')+']\0'+'---Not converged---')
                shutil.move(workpath+"\\"+name,workpath+"\\"+name+"Not converged")
                sheet.append([name, 'Not converged.']) 



            #exit fluent
            coWbUnit.execWbCommand('setup1.Exit()')
            print('['+datetime.datetime.now().strftime('%F %T')+']\0'+name+'\0Fluent Completed.')
            wb.save(workpath+"\\caculate_dp.xlsx")

        except Exception as e1:
            print()
            print(e1)
            print('['+datetime.datetime.now().strftime('%F %T')+'] Fluent exception!')
            print()

        finally:
            #Save
            coWbUnit.saveProject(workpath+"\\"+name+"\\"+wbpjname)
            coWbUnit.finalize()
            print('['+datetime.datetime.now().strftime('%F %T')+']\0'+name+'\0Save Completed.')
            print()    
            t1=time.time()
            deltat=int(t1)-int(t0) 
            ET=datetime.timedelta(seconds=deltat)
            ETA=datetime.timedelta(seconds=deltat*(int(n)-int(i)-1))
            percentage=round(((i+1)/int(n))*100,2)
            print('\0'+str(percentage)+' % ( '+str(i+1)+' of '+n+' ) [ Elapsed Time: '+str(ET)+' ] [ ETA: '+str(ETA)+' ]')
            i+=1
    except Exception as e2:
        print()
        print(e2)
        print('['+datetime.datetime.now().strftime('%F %T')+'] Program exception!')
        print()
        i+=1


print()
print('====================================================================================')
print()
print('['+datetime.datetime.now().strftime('%F %T')+'] Work is completed.')
print()
print('====================================================================================')

#exit
replacement(Fluentscriptname, MaterialCommand, var1) #change the script to origin
replacement(Fluentscriptname, MassFlowCommand, var2) #change the script to origin
print()
print('Press any key to continue...')
input()
print("The program will terminate in ten seconds.")
time.sleep(10)
sys.exit()
