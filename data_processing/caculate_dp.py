import linecache
import os
import sys
import time
import openpyxl #third package

#count the total lines
def count_lines(file):
    count = -1
    for count, line in enumerate(open(file, 'rb')):
        pass
        count += 1
    return count

#if last_phead <±5 than the previous one. Compare 10 dp.
def compare_phead(file):
    count = count_lines(file)
    i=0
    p=[0 for x in range(0, 20)]
    while(i<20):
        text1=linecache.getline(file,count-i).split()
        p[i]=float(text1[1])
        i+=1
    m=0
    j=0
    while(j<10):
        k=0
        sum1=sum2=0
        while(j+k<=j+9):
            sum1+=p[k+j]
            sum2+=p[k+j+1]
            k+=1
        avg_dp=(sum1-sum2)/10
        j+=1
        if(avg_dp >= -5 and avg_dp <= 5):
            m+=1
    if(m==10):
        result=(sum1+sum2)/20
        return result
    else:
        return '0'

workpath = sys.path[0]  

wb=openpyxl.Workbook() #create Workbook() object
ws=wb.active #get dafault sheet
ws.append(["Name", "p_head(Pa)","p_head(mmHg)"])#write data to file
wb.save("caculate_dp.xlsx")

filePath=os.listdir(workpath)
i=0
while(i<len(filePath)-2):
    name=filePath[i]
    fname_p_head = workpath+"\\"+name+"\\"+name+"_files"+"\\dp0\\FFF\\Fluent\\p-head-rfile.out"
    wb=openpyxl.load_workbook("caculate_dp.xlsx")
    sheet=wb['Sheet'] #Get sheet by name
    
    print('====================================================================================')
    print()
    print(name)
    cp=compare_phead(fname_p_head)
    if(cp!='0'):
        print(' p-head= {:.4e} Pa'.format(float(cp)))
        print('       = {:.4e} mmHg'.format(float(cp)*0.0075))
        print()
        sheet.append([name, round(float(cp),4),round(float(cp)*0.0075,4)]) 
    else:
        print('Not converged.')
        print()
        sheet.append([name, 'Not converged.']) 
    wb.save("caculate_dp.xlsx")
    i+=1

print('====================================================================================')

#exit
print()
print('Press any key to continue...')
input()
print("The program will terminate in ten seconds.")
time.sleep(10)
sys.exit()
